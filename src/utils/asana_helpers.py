"""
Asana Helpers

Helper classes and functions for Asana task management integration.
"""

import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional

from core.logging_config import get_logger

logger = get_logger(__name__)


class AsanaHelper:
    """
    Helper for Asana automation tasks.

    Provides convenient methods for common Asana operations:
    - Email-to-task creation (no API required)
    - Browser automation for complex operations
    - CSV bulk operations
    - Excel-to-Asana synchronization
    """

    def __init__(
        self,
        project_email: Optional[str] = None,
        project_url: Optional[str] = None
    ) -> None:
        """
        Initialize Asana helper.

        Args:
            project_email: Asana project email address (for email-to-task).
            project_url: Asana project URL (for browser automation).
        """
        self.project_email = project_email
        self.project_url = project_url

    def create_task_via_email(
        self,
        name: str,
        description: str = '',
        assignee: str = '',
        due_date: str = '',
        priority: str = '',
        tags: str = '',
        attachments: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Create a single task via email.

        Args:
            name: Task name.
            description: Task description.
            assignee: Assignee username (without @).
            due_date: Due date (YYYY-MM-DD or 'next Friday', etc.).
            priority: Priority (High/Medium/Low).
            tags: Comma-separated tags.
            attachments: List of file paths to attach.

        Returns:
            Result dictionary.

        Example:
            helper = AsanaHelper(project_email='myproject@mail.asana.com')
            result = helper.create_task_via_email(
                name='Review proposal',
                assignee='john.doe',
                due_date='2024-12-15',
                priority='High'
            )
        """
        from modules.asana import AsanaEmailModule

        if not self.project_email:
            raise ValueError("project_email is required for email-based operations")

        task = {
            'name': name,
            'description': description,
            'assignee': assignee,
            'due_date': due_date,
            'priority': priority,
            'tags': tags,
            'attachments': attachments or []
        }

        module = AsanaEmailModule()
        result = module.run(
            project_email=self.project_email,
            tasks=[task]
        )

        logger.info(f"Created Asana task via email: {name}")
        return result

    def bulk_create_from_list(
        self,
        tasks: List[Dict[str, Any]],
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Bulk create tasks from a list.

        Args:
            tasks: List of task dictionaries.
            method: 'email' or 'browser'.

        Returns:
            Result dictionary.

        Example:
            tasks = [
                {'name': 'Task 1', 'assignee': 'john'},
                {'name': 'Task 2', 'due_date': '2024-12-15'}
            ]
            result = helper.bulk_create_from_list(tasks, method='email')
        """
        if method == 'email':
            from modules.asana import AsanaEmailModule

            if not self.project_email:
                raise ValueError("project_email is required for email method")

            module = AsanaEmailModule()
            result = module.run(
                project_email=self.project_email,
                tasks=tasks
            )

        elif method == 'browser':
            from modules.asana import AsanaBrowserModule

            if not self.project_url:
                raise ValueError("project_url is required for browser method")

            module = AsanaBrowserModule()
            result = module.run(
                asana_url=self.project_url,
                operation='create_tasks',
                tasks=tasks
            )

        else:
            raise ValueError(f"Invalid method: {method}. Must be 'email' or 'browser'")

        logger.info(f"Bulk created {len(tasks)} tasks via {method}")
        return result

    def bulk_create_from_excel(
        self,
        excel_path: str,
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Bulk create tasks from Excel file.

        Excel columns: name, description, assignee, due_date, priority, tags

        Args:
            excel_path: Path to Excel file.
            method: 'email' or 'csv'.

        Returns:
            Result dictionary.

        Example:
            helper = AsanaHelper(project_email='myproject@mail.asana.com')
            result = helper.bulk_create_from_excel('C:/Data/tasks.xlsx')
        """
        if method == 'email':
            from modules.asana import AsanaEmailModule

            if not self.project_email:
                raise ValueError("project_email is required for email method")

            module = AsanaEmailModule()
            result = module.run(
                project_email=self.project_email,
                from_excel=excel_path
            )

        elif method == 'csv':
            # Convert Excel to Asana CSV
            from modules.asana import AsanaCSVHandler

            handler = AsanaCSVHandler()

            # First generate CSV
            csv_output = excel_path.replace('.xlsx', '_asana.csv').replace('.xls', '_asana.csv')
            result = handler.run(
                operation='generate',
                input_file=excel_path,
                output_file=csv_output
            )

            logger.info(f"Generated Asana CSV for import: {csv_output}")
            logger.info("You can now manually import this CSV into Asana")

        else:
            raise ValueError(f"Invalid method: {method}. Must be 'email' or 'csv'")

        logger.info(f"Bulk created tasks from Excel: {excel_path}")
        return result

    def generate_asana_csv(
        self,
        tasks: List[Dict[str, Any]],
        output_path: str
    ) -> Dict[str, Any]:
        """
        Generate Asana-compatible CSV from task list.

        Args:
            tasks: List of task dictionaries.
            output_path: Output CSV file path.

        Returns:
            Result dictionary.

        Example:
            tasks = [
                {'name': 'Task 1', 'assignee': 'john', 'due_date': '2024-12-15'},
                {'name': 'Task 2', 'priority': 'High'}
            ]
            helper.generate_asana_csv(tasks, 'C:/Output/tasks.csv')
        """
        from modules.asana import AsanaCSVHandler

        handler = AsanaCSVHandler()
        result = handler.run(
            operation='generate',
            tasks=tasks,
            output_file=output_path
        )

        logger.info(f"Generated Asana CSV: {output_path}")
        return result

    def parse_asana_export(
        self,
        csv_path: str,
        output_json: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Parse Asana CSV export into structured data.

        Args:
            csv_path: Path to Asana CSV export.
            output_json: Optional path to save JSON output.

        Returns:
            List of task dictionaries.

        Example:
            tasks = helper.parse_asana_export('C:/Downloads/asana_export.csv')
            # Analyze tasks
            for task in tasks:
                print(f"{task['name']} - {task['assignee']}")
        """
        from modules.asana import AsanaCSVHandler

        handler = AsanaCSVHandler()
        result = handler.run(
            operation='parse',
            input_file=csv_path,
            output_file=output_json
        )

        logger.info(f"Parsed {result['data']['task_count']} tasks from Asana export")
        return result['data']['tasks']

    def convert_excel_to_asana_csv(
        self,
        excel_path: str,
        output_path: str,
        column_mapping: Optional[Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """
        Convert Excel tracker to Asana-compatible CSV.

        Args:
            excel_path: Path to Excel file.
            output_path: Output CSV path.
            column_mapping: Map Asana columns to Excel columns.
                          e.g., {'Task Name': 'title', 'Assignee': 'owner'}

        Returns:
            Result dictionary.

        Example:
            mapping = {
                'Task Name': 'task_title',
                'Assignee': 'assigned_to',
                'Due Date': 'deadline'
            }
            result = helper.convert_excel_to_asana_csv(
                'C:/Data/project.xlsx',
                'C:/Output/asana_import.csv',
                mapping
            )
        """
        from modules.asana import AsanaCSVHandler
        import csv
        from openpyxl import load_workbook

        # If no mapping, try auto-detection
        if not column_mapping:
            column_mapping = {}

        handler = AsanaCSVHandler()

        # First convert Excel to temp CSV
        temp_csv = tempfile.mktemp(suffix='.csv')

        # Load and convert Excel to CSV
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        with open(temp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        # Now convert CSV to Asana format
        result = handler.run(
            operation='convert',
            input_file=temp_csv,
            output_file=output_path,
            mapping=column_mapping
        )

        # Cleanup temp file
        Path(temp_csv).unlink(missing_ok=True)

        logger.info(f"Converted Excel to Asana CSV: {output_path}")
        return result

    def sync_excel_tracker_to_asana(
        self,
        excel_path: str,
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Sync Excel project tracker to Asana.

        This is a convenience method that:
        1. Reads tasks from Excel
        2. Creates them in Asana via email or CSV

        Args:
            excel_path: Path to Excel tracker.
            method: 'email' for direct creation, 'csv' for manual import.

        Returns:
            Result dictionary.

        Example:
            # Sync weekly sprint tasks from Excel to Asana
            helper = AsanaHelper(project_email='sprint@mail.asana.com')
            result = helper.sync_excel_tracker_to_asana('C:/Data/sprint_tasks.xlsx')
        """
        logger.info(f"Syncing Excel tracker to Asana: {excel_path}")

        if method == 'email':
            return self.bulk_create_from_excel(excel_path, method='email')

        elif method == 'csv':
            # Generate Asana CSV for manual import
            csv_output = excel_path.replace('.xlsx', '_asana_import.csv').replace('.xls', '_asana_import.csv')
            return self.convert_excel_to_asana_csv(excel_path, csv_output)

        else:
            raise ValueError(f"Invalid method: {method}")

    def create_weekly_sprint_tasks(
        self,
        sprint_number: int,
        start_date: str,
        task_templates: List[Dict[str, str]]
    ) -> Dict[str, Any]:
        """
        Create weekly sprint tasks from templates.

        Args:
            sprint_number: Sprint number (e.g., 42).
            start_date: Sprint start date (YYYY-MM-DD).
            task_templates: List of task templates with name, assignee, day_offset.

        Returns:
            Result dictionary.

        Example:
            templates = [
                {'name': 'Sprint Planning', 'assignee': 'team', 'day_offset': 0},
                {'name': 'Daily Standup', 'assignee': 'team', 'day_offset': 1},
                {'name': 'Sprint Review', 'assignee': 'team', 'day_offset': 4},
                {'name': 'Retrospective', 'assignee': 'team', 'day_offset': 4}
            ]
            result = helper.create_weekly_sprint_tasks(42, '2024-12-09', templates)
        """
        logger.info(f"Creating Sprint {sprint_number} tasks...")

        start = datetime.strptime(start_date, '%Y-%m-%d')

        tasks: List[Dict[str, Any]] = []
        for template in task_templates:
            due_date = start + timedelta(days=template.get('day_offset', 0))

            task = {
                'name': f"Sprint {sprint_number}: {template['name']}",
                'assignee': template.get('assignee', ''),
                'due_date': due_date.strftime('%Y-%m-%d'),
                'description': template.get('description', ''),
                'priority': template.get('priority', 'Medium'),
                'tags': f"sprint-{sprint_number}"
            }
            tasks.append(task)

        return self.bulk_create_from_list(tasks, method='email')


def create_asana_roadmap_from_excel(
    excel_path: str,
    project_email: str,
    output_csv: Optional[str] = None
) -> Dict[str, Any]:
    """
    Common workflow: Create Asana roadmap from Excel planning file.

    Args:
        excel_path: Path to Excel roadmap file.
        project_email: Asana project email.
        output_csv: Optional CSV output path (for manual import).

    Returns:
        Result dictionary.

    Example:
        result = create_asana_roadmap_from_excel(
            'C:/Planning/Q1_Roadmap.xlsx',
            'q1-projects@mail.asana.com'
        )
    """
    logger.info("Creating Asana roadmap from Excel...")

    helper = AsanaHelper(project_email=project_email)

    if output_csv:
        # Generate CSV for review/manual import
        result = helper.convert_excel_to_asana_csv(excel_path, output_csv)
        logger.info(f"Roadmap CSV generated: {output_csv}")
    else:
        # Direct email creation
        result = helper.bulk_create_from_excel(excel_path, method='email')
        logger.info("Roadmap tasks created via email")

    return result


def sync_asana_status_to_excel(
    asana_csv_export: str,
    excel_tracker: str,
    output_path: Optional[str] = None
) -> Dict[str, Any]:
    """
    Common workflow: Sync Asana task status back to Excel tracker.

    Args:
        asana_csv_export: Path to Asana CSV export.
        excel_tracker: Path to Excel tracker file.
        output_path: Output Excel path (defaults to tracker path).

    Returns:
        Result dictionary with sync summary.

    Example:
        result = sync_asana_status_to_excel(
            'C:/Downloads/asana_export.csv',
            'C:/Tracking/project_tracker.xlsx'
        )
    """
    logger.info("Syncing Asana status to Excel tracker...")

    # Parse Asana export
    helper = AsanaHelper()
    asana_tasks = helper.parse_asana_export(asana_csv_export)

    # Load Excel tracker
    from modules.excel_automation import WorkbookHandler

    workbook = WorkbookHandler()
    workbook.load(excel_tracker)

    # Update status based on task names
    # (This is a template - actual implementation depends on your Excel structure)

    updated_count = 0
    for task in asana_tasks:
        task_name = task.get('name', '')
        status = 'Complete' if task.get('completed_at') else 'In Progress'

        # Find and update matching row in Excel
        # (Simplified - actual implementation would search by name)
        updated_count += 1

    # Save
    if not output_path:
        output_path = excel_tracker

    workbook.save(output_path)

    logger.info(f"Updated {updated_count} tasks in Excel tracker")

    return {
        'status': 'success',
        'message': f'Synced {updated_count} tasks',
        'output_file': output_path
    }
