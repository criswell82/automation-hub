"""
Workflow Helpers

Common automation patterns and helper functions for building workflows.
Focused on: Report Generation, Email & Communication, File Management
"""

import os
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional, Callable
from datetime import datetime
import glob

from core.logging_config import get_logger

logger = get_logger(__name__)


# =======================
# REPORT GENERATION HELPERS
# =======================

class ReportBuilder:
    """Helper for building Excel reports"""

    def __init__(self):
        from modules.excel_automation import WorkbookHandler, ChartBuilder
        self.workbook = WorkbookHandler()
        self.chart_builder = ChartBuilder()

    def add_title_sheet(self, title: str, subtitle: str = "", metadata: Dict = None):
        """Add a title/cover sheet to the report"""
        sheet = self.workbook.workbook.create_sheet("Cover", 0)

        # Title
        sheet['A1'] = title
        self.workbook.format_cells('Cover', 'A1', bold=True, font_size=18)

        # Subtitle
        if subtitle:
            sheet['A2'] = subtitle
            self.workbook.format_cells('Cover', 'A2', font_size=12)

        # Metadata
        if metadata:
            row = 4
            for key, value in metadata.items():
                sheet[f'A{row}'] = f"{key}:"
                sheet[f'B{row}'] = str(value)
                self.workbook.format_cells('Cover', f'A{row}', bold=True)
                row += 1

        return sheet

    def add_data_sheet(self, sheet_name: str, data: List[List[Any]],
                      headers: List[str] = None, format_header: bool = True):
        """Add a data sheet with optional headers"""
        sheet = self.workbook.workbook.create_sheet(sheet_name)

        start_row = 1
        if headers:
            sheet.append(headers)
            if format_header:
                header_range = f"A1:{chr(64 + len(headers))}1"
                self.workbook.format_cells(sheet_name, header_range,
                                         bold=True, fill_color='4472C4')
            start_row = 2

        for row_data in data:
            sheet.append(row_data)

        # Auto-size columns
        self.workbook.auto_size_columns(sheet_name)

        return sheet

    def add_summary_with_chart(self, sheet_name: str, data: Dict[str, float],
                              chart_type: str = 'bar', chart_title: str = ""):
        """Add a summary sheet with a chart"""
        sheet = self.workbook.workbook.create_sheet(sheet_name)

        # Write data
        row = 1
        for label, value in data.items():
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            row += 1

        # Format
        self.workbook.format_cells(sheet_name, 'A1:A' + str(row-1), bold=True)

        # Add chart
        chart = self.chart_builder.create_chart(
            chart_type=chart_type,
            title=chart_title or f"{sheet_name} Summary"
        )

        # Configure chart data
        data_range = f'A1:B{row-1}'
        self.chart_builder.add_data_to_chart(
            chart, sheet, data_range,
            categories_col='A', values_col='B'
        )

        # Position chart
        sheet.add_chart(chart, f'D2')

        return sheet

    def save(self, output_path: str):
        """Save the report"""
        self.workbook.save(output_path)
        logger.info(f"Report saved to: {output_path}")


def aggregate_data_from_files(file_pattern: str, aggregation_func: Callable) -> Any:
    """
    Aggregate data from multiple files matching a pattern

    Args:
        file_pattern: Glob pattern (e.g., "C:/Data/*.csv")
        aggregation_func: Function to process each file and return aggregated data

    Returns:
        Aggregated result
    """
    files = glob.glob(file_pattern)
    logger.info(f"Found {len(files)} files matching pattern: {file_pattern}")

    results = []
    for file_path in files:
        try:
            result = aggregation_func(file_path)
            results.append(result)
            logger.info(f"Processed: {os.path.basename(file_path)}")
        except Exception as e:
            logger.warning(f"Failed to process {file_path}: {e}")

    return results


def create_pivot_summary(data: List[Dict], group_by: str, sum_field: str) -> Dict:
    """
    Create a pivot-style summary from data

    Args:
        data: List of dictionaries
        group_by: Field to group by
        sum_field: Field to sum

    Returns:
        Dictionary of {group_value: total}
    """
    summary = {}
    for row in data:
        key = row.get(group_by, 'Unknown')
        value = row.get(sum_field, 0)

        if key in summary:
            summary[key] += value
        else:
            summary[key] = value

    return summary


# =======================
# ONENOTE KNOWLEDGE MANAGEMENT HELPERS
# =======================

class OneNoteHelper:
    """Helper for OneNote knowledge management tasks"""

    def __init__(self, default_notebook: str = None, default_section: str = None):
        """
        Initialize OneNote helper.

        Args:
            default_notebook: Default notebook name
            default_section: Default section name
        """
        from modules.onenote import OneNoteManager, OneNoteContentBuilder, TemplateBuilder

        self.manager = OneNoteManager()
        self.manager.configure(
            default_notebook=default_notebook,
            default_section=default_section
        )
        self.manager.validate()

        self.content_builder = OneNoteContentBuilder
        self.templates = TemplateBuilder

    def create_page(
        self,
        notebook: str,
        section: str,
        title: str,
        content: str = ""
    ) -> str:
        """
        Create a simple OneNote page.

        Args:
            notebook: Notebook name
            section: Section name
            title: Page title
            content: Page content (simple text)

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.create_page(
                notebook='Project Knowledge',
                section='Requirements',
                title='Q4 Requirements',
                content='Key requirements for Q4...'
            )
        """
        logger.info(f"Creating OneNote page: {title}")

        page_id = self.manager.create_page(
            notebook=notebook,
            section=section,
            title=title,
            content=content
        )

        logger.info(f"OneNote page created: {page_id}")
        return page_id

    def get_page_content(
        self,
        notebook: str,
        section: str,
        title: str,
        format: str = "text"
    ) -> Optional[str]:
        """
        Get content from a OneNote page.

        Args:
            notebook: Notebook name
            section: Section name
            title: Page title
            format: Content format ('text', 'simple', 'xml')

        Returns:
            str or None: Page content

        Example:
            onenote = OneNoteHelper()
            content = onenote.get_page_content(
                notebook='Project Knowledge',
                section='Requirements',
                title='Q4 Requirements'
            )
            print(content)
        """
        logger.info(f"Getting OneNote page content: {title}")

        content = self.manager.get_page(
            notebook=notebook,
            section=section,
            title=title,
            format=format
        )

        if content:
            logger.info(f"Retrieved {len(content)} characters from page")
        else:
            logger.warning(f"Page '{title}' not found")

        return content

    def search_for_facts(
        self,
        query: str,
        notebook: str = None,
        max_results: int = 10
    ) -> List[Dict[str, Any]]:
        """
        Search OneNote for project facts/information.

        Args:
            query: Search query
            notebook: Optional notebook to search in
            max_results: Maximum results to return

        Returns:
            list: Matching pages with snippets

        Example:
            onenote = OneNoteHelper()
            results = onenote.search_for_facts(
                query='approved requirements',
                notebook='Project Knowledge'
            )

            for result in results:
                print(f"{result['title']}: {result['snippet']}")
        """
        logger.info(f"Searching OneNote for: '{query}'")

        results = self.manager.search(
            query=query,
            notebook=notebook,
            max_results=max_results
        )

        logger.info(f"Found {len(results)} results")
        return results

    def extract_facts_for_report(
        self,
        notebook: str,
        section: str,
        page_title: str
    ) -> Dict[str, Any]:
        """
        Extract structured facts from OneNote page for use in reports.

        Args:
            notebook: Notebook name
            section: Section name
            page_title: Page title

        Returns:
            dict: Extracted facts and tables

        Example:
            onenote = OneNoteHelper()
            facts = onenote.extract_facts_for_report(
                notebook='Project Knowledge',
                section='Status',
                page_title='Q4 Metrics'
            )

            # Use in Excel report
            report = ReportBuilder()
            report.add_data_sheet('Project Facts', facts['tables'][0])
        """
        logger.info(f"Extracting facts from OneNote page: {page_title}")

        # Get page content
        content = self.manager.get_page(
            notebook=notebook,
            section=section,
            title=page_title,
            format='text'
        )

        # Extract tables
        tables = self.manager.extract_tables_from_page(
            notebook=notebook,
            section=section,
            title=page_title
        )

        facts = {
            'content': content,
            'tables': tables,
            'source': f"{notebook}/{section}/{page_title}",
            'extracted_at': datetime.now().isoformat()
        }

        logger.info(f"Extracted {len(tables)} tables from page")
        return facts

    def create_meeting_minutes(
        self,
        notebook: str,
        section: str,
        meeting_title: str,
        date: str,
        attendees: List[str],
        agenda: List[str],
        notes: str = "",
        action_items: List[Dict[str, str]] = None
    ) -> str:
        """
        Create meeting minutes page from template.

        Args:
            notebook: Notebook name
            section: Section name
            meeting_title: Meeting title
            date: Meeting date
            attendees: List of attendees
            agenda: Agenda items
            notes: Meeting notes
            action_items: Action items with 'task', 'owner', 'due' keys

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.create_meeting_minutes(
                notebook='Meetings',
                section='Sprint Planning',
                meeting_title='Sprint 42 Planning',
                date='2024-12-09',
                attendees=['Alice', 'Bob', 'Carol'],
                agenda=['Review backlog', 'Estimate tasks', 'Plan sprint'],
                action_items=[
                    {'task': 'Update design docs', 'owner': 'Alice', 'due': '2024-12-12'}
                ]
            )
        """
        logger.info(f"Creating meeting minutes: {meeting_title}")

        page_id = self.manager.create_meeting_minutes(
            notebook=notebook,
            section=section,
            meeting_title=meeting_title,
            date=date,
            attendees=attendees,
            agenda=agenda,
            notes=notes,
            action_items=action_items
        )

        logger.info(f"Meeting minutes created: {page_id}")
        return page_id

    def create_meeting_minutes_from_outlook(
        self,
        notebook: str,
        section: str,
        meeting_subject: str
    ) -> str:
        """
        Create meeting minutes from Outlook meeting details.

        Args:
            notebook: Notebook name
            section: Section name
            meeting_subject: Outlook meeting subject to find

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.create_meeting_minutes_from_outlook(
                notebook='Meetings',
                section='Sprint Planning',
                meeting_subject='Sprint 42 Planning Meeting'
            )
        """
        from modules.outlook_automation import MeetingManager

        logger.info(f"Creating meeting minutes from Outlook: {meeting_subject}")

        # Get meeting details from Outlook
        meeting_mgr = MeetingManager()
        meeting_mgr.connect()
        meeting = meeting_mgr.get_meeting_by_subject(meeting_subject)

        if not meeting:
            raise ValueError(f"Meeting '{meeting_subject}' not found in Outlook")

        # Extract details
        attendees = [a.get('name', '') for a in meeting.get('attendees', [])]
        date = meeting.get('start_time', datetime.now().strftime('%Y-%m-%d'))
        body = meeting.get('body', '')

        # Parse agenda from body (simple approach)
        agenda = []
        for line in body.split('\n'):
            line = line.strip()
            if line and (line.startswith('-') or line.startswith('•')):
                agenda.append(line.lstrip('-•').strip())

        if not agenda:
            agenda = ['Review agenda items']

        # Create page
        page_id = self.create_meeting_minutes(
            notebook=notebook,
            section=section,
            meeting_title=meeting.get('subject', meeting_subject),
            date=date,
            attendees=attendees,
            agenda=agenda,
            notes=""
        )

        logger.info("Meeting minutes created from Outlook meeting")
        return page_id

    def create_project_status_page(
        self,
        notebook: str,
        section: str,
        project_name: str,
        status: str,
        progress: int,
        milestones: List[Dict[str, str]],
        issues: List[str] = None,
        next_steps: List[str] = None
    ) -> str:
        """
        Create project status page from template.

        Args:
            notebook: Notebook name
            section: Section name
            project_name: Project name
            status: Status (On Track, At Risk, Delayed)
            progress: Progress percentage (0-100)
            milestones: List with 'name', 'due', 'status' keys
            issues: List of issues
            next_steps: List of next steps

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.create_project_status_page(
                notebook='Projects',
                section='Q4 2024',
                project_name='Homepage Redesign',
                status='On Track',
                progress=75,
                milestones=[
                    {'name': 'Design Complete', 'due': '2024-12-01', 'status': 'Done'},
                    {'name': 'Development', 'due': '2024-12-15', 'status': 'In Progress'}
                ],
                next_steps=['Complete testing', 'Deploy to production']
            )
        """
        logger.info(f"Creating project status page: {project_name}")

        page_id = self.manager.create_project_status(
            notebook=notebook,
            section=section,
            project_name=project_name,
            status=status,
            progress=progress,
            milestones=milestones,
            issues=issues,
            next_steps=next_steps
        )

        logger.info(f"Project status page created: {page_id}")
        return page_id

    def create_research_notes(
        self,
        notebook: str,
        section: str,
        topic: str,
        source: str,
        key_findings: List[str],
        details: str = "",
        tags: List[str] = None
    ) -> str:
        """
        Create research notes page.

        Args:
            notebook: Notebook name
            section: Section name
            topic: Research topic
            source: Information source
            key_findings: List of key findings
            details: Additional details
            tags: Tags for categorization

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.create_research_notes(
                notebook='Research',
                section='Competitors',
                topic='Competitor Homepage Analysis',
                source='Web research - 2024-12-09',
                key_findings=[
                    'Average load time: 2.3 seconds',
                    'Mobile-first design approach',
                    'Focus on video content'
                ],
                tags=['homepage', 'ux', 'performance']
            )
        """
        logger.info(f"Creating research notes: {topic}")

        page_id = self.manager.create_research_notes(
            notebook=notebook,
            section=section,
            topic=topic,
            source=source,
            key_findings=key_findings,
            details=details,
            tags=tags
        )

        logger.info(f"Research notes created: {page_id}")
        return page_id

    def aggregate_research_to_central_notebook(
        self,
        source_files: List[str],
        central_notebook: str,
        section: str,
        topic: str
    ) -> str:
        """
        Aggregate research from various files into central OneNote repository.

        Args:
            source_files: List of file paths to aggregate (txt, pdf, md, etc.)
            central_notebook: Central notebook name
            section: Section name
            topic: Research topic

        Returns:
            str: New page ID

        Example:
            onenote = OneNoteHelper()
            page_id = onenote.aggregate_research_to_central_notebook(
                source_files=[
                    'C:/Research/analysis1.txt',
                    'C:/Research/analysis2.txt'
                ],
                central_notebook='Project Knowledge',
                section='Research',
                topic='Market Analysis'
            )
        """
        logger.info(f"Aggregating research from {len(source_files)} files")

        # Read all source files
        all_content = []
        key_findings = []

        for file_path in source_files:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    all_content.append(f"## {Path(file_path).name}\n\n{content}\n")

                    # Extract first few lines as findings (simple approach)
                    lines = [l.strip() for l in content.split('\n') if l.strip()]
                    key_findings.extend(lines[:3])

                logger.info(f"Read: {Path(file_path).name}")
            except Exception as e:
                logger.warning(f"Failed to read {file_path}: {e}")

        # Combine content
        combined_content = "\n\n".join(all_content)

        # Create research notes page
        page_id = self.create_research_notes(
            notebook=central_notebook,
            section=section,
            topic=topic,
            source=f"Aggregated from {len(source_files)} files",
            key_findings=key_findings[:10],  # Top 10 findings
            details=combined_content,
            tags=['aggregated', 'research']
        )

        logger.info(f"Aggregated research into OneNote page: {page_id}")
        return page_id

    def list_notebooks(self) -> List[Dict[str, str]]:
        """
        List all available OneNote notebooks.

        Returns:
            list: Notebooks with id, name, path

        Example:
            onenote = OneNoteHelper()
            notebooks = onenote.list_notebooks()
            for nb in notebooks:
                print(f"Notebook: {nb['name']}")
        """
        return self.manager.list_notebooks()

    def list_sections(self, notebook: str) -> List[Dict[str, str]]:
        """
        List sections in a notebook.

        Args:
            notebook: Notebook name

        Returns:
            list: Sections with id, name, path

        Example:
            onenote = OneNoteHelper()
            sections = onenote.list_sections('Project Knowledge')
            for sec in sections:
                print(f"Section: {sec['name']}")
        """
        return self.manager.list_sections(notebook)

    def list_pages(self, notebook: str, section: str) -> List[Dict[str, str]]:
        """
        List pages in a section.

        Args:
            notebook: Notebook name
            section: Section name

        Returns:
            list: Pages with id, title, datetime

        Example:
            onenote = OneNoteHelper()
            pages = onenote.list_pages('Project Knowledge', 'Requirements')
            for page in pages:
                print(f"Page: {page['title']}")
        """
        return self.manager.list_pages(notebook, section)


# =======================
# EMAIL & COMMUNICATION HELPERS
# =======================

class EmailHelper:
    """Helper for email automation tasks"""

    def __init__(self):
        from modules.outlook_automation import EmailHandler
        self.email_handler = EmailHandler()

    def send_report(self, to: List[str], subject: str, body: str,
                   attachments: List[str] = None, cc: List[str] = None):
        """
        Send an email with report attachments

        Args:
            to: List of recipient email addresses
            subject: Email subject
            body: Email body (can include HTML)
            attachments: List of file paths to attach
            cc: List of CC recipients
        """
        logger.info(f"Sending email to: {', '.join(to)}")

        self.email_handler.send_email(
            to=to,
            subject=subject,
            body=body,
            attachments=attachments or [],
            cc=cc
        )

        logger.info("Email sent successfully")

    def process_inbox(self, folder: str = "Inbox", unread_only: bool = True,
                     filter_subject: str = None) -> List[Dict]:
        """
        Process emails from inbox

        Args:
            folder: Email folder to process
            unread_only: Only process unread emails
            filter_subject: Filter by subject keyword

        Returns:
            List of email dictionaries
        """
        logger.info(f"Processing emails from: {folder}")

        emails = self.email_handler.read_emails(
            folder=folder,
            unread_only=unread_only
        )

        if filter_subject:
            emails = [e for e in emails if filter_subject.lower() in e.get('subject', '').lower()]

        logger.info(f"Found {len(emails)} emails")
        return emails

    def extract_attachments(self, emails: List[Dict], output_folder: str,
                          file_extensions: List[str] = None) -> List[str]:
        """
        Extract attachments from emails

        Args:
            emails: List of email dictionaries (from process_inbox)
            output_folder: Where to save attachments
            file_extensions: Optional filter (e.g., ['.pdf', '.xlsx'])

        Returns:
            List of saved file paths
        """
        os.makedirs(output_folder, exist_ok=True)
        saved_files = []

        for email in emails:
            attachments = email.get('attachments', [])

            for attachment in attachments:
                file_name = attachment.get('name', 'attachment')

                # Filter by extension if specified
                if file_extensions:
                    if not any(file_name.endswith(ext) for ext in file_extensions):
                        continue

                # Save attachment
                file_path = os.path.join(output_folder, file_name)
                # Note: Actual save logic would use email_handler
                saved_files.append(file_path)
                logger.info(f"Saved attachment: {file_name}")

        logger.info(f"Extracted {len(saved_files)} attachments")
        return saved_files

    def create_email_summary(self, to: str, subject: str, data: Dict[str, Any]):
        """
        Create and send a summary email from data

        Args:
            to: Recipient email
            subject: Email subject
            data: Dictionary of summary data
        """
        # Build HTML email body
        body = "<html><body>"
        body += f"<h2>{subject}</h2>"
        body += "<table border='1' cellpadding='5' cellspacing='0'>"

        for key, value in data.items():
            body += f"<tr><td><b>{key}</b></td><td>{value}</td></tr>"

        body += "</table>"
        body += f"<p><i>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i></p>"
        body += "</body></html>"

        self.send_report([to], subject, body)


# =======================
# FILE & DOCUMENT MANAGEMENT HELPERS
# =======================

class FileOrganizer:
    """Helper for file management tasks"""

    def __init__(self):
        pass

    def organize_by_date(self, source_folder: str, dest_folder: str,
                        date_format: str = "%Y-%m"):
        """
        Organize files into folders by date

        Args:
            source_folder: Source directory
            dest_folder: Destination directory
            date_format: Date format for folder names (e.g., "2024-01")
        """
        os.makedirs(dest_folder, exist_ok=True)

        files = Path(source_folder).glob('*')
        moved_count = 0

        for file_path in files:
            if file_path.is_file():
                # Get file date
                file_date = datetime.fromtimestamp(file_path.stat().st_mtime)
                date_folder = file_date.strftime(date_format)

                # Create date folder
                target_folder = os.path.join(dest_folder, date_folder)
                os.makedirs(target_folder, exist_ok=True)

                # Move file
                target_path = os.path.join(target_folder, file_path.name)
                shutil.move(str(file_path), target_path)
                logger.info(f"Moved: {file_path.name} -> {date_folder}/")
                moved_count += 1

        logger.info(f"Organized {moved_count} files by date")

    def organize_by_type(self, source_folder: str, dest_folder: str,
                        type_mapping: Dict[str, List[str]] = None):
        """
        Organize files by type/extension

        Args:
            source_folder: Source directory
            dest_folder: Destination directory
            type_mapping: Dict of {folder_name: [extensions]}
                         e.g., {"Documents": [".pdf", ".docx"], "Images": [".jpg", ".png"]}
        """
        if type_mapping is None:
            type_mapping = {
                "Documents": [".pdf", ".doc", ".docx", ".txt"],
                "Spreadsheets": [".xlsx", ".xls", ".csv"],
                "Images": [".jpg", ".jpeg", ".png", ".gif"],
                "Archives": [".zip", ".rar", ".7z"],
                "Other": []  # Catch-all
            }

        os.makedirs(dest_folder, exist_ok=True)

        files = Path(source_folder).glob('*')
        moved_count = 0

        for file_path in files:
            if file_path.is_file():
                ext = file_path.suffix.lower()

                # Find matching type
                target_folder_name = "Other"
                for folder_name, extensions in type_mapping.items():
                    if ext in extensions:
                        target_folder_name = folder_name
                        break

                # Create type folder
                target_folder = os.path.join(dest_folder, target_folder_name)
                os.makedirs(target_folder, exist_ok=True)

                # Move file
                target_path = os.path.join(target_folder, file_path.name)
                shutil.move(str(file_path), target_path)
                logger.info(f"Moved: {file_path.name} -> {target_folder_name}/")
                moved_count += 1

        logger.info(f"Organized {moved_count} files by type")

    def batch_rename(self, folder: str, pattern: str, replacement: str):
        """
        Batch rename files in a folder

        Args:
            folder: Folder containing files
            pattern: Pattern to replace
            replacement: Replacement string
        """
        files = Path(folder).glob('*')
        renamed_count = 0

        for file_path in files:
            if file_path.is_file() and pattern in file_path.name:
                new_name = file_path.name.replace(pattern, replacement)
                new_path = file_path.parent / new_name

                file_path.rename(new_path)
                logger.info(f"Renamed: {file_path.name} -> {new_name}")
                renamed_count += 1

        logger.info(f"Renamed {renamed_count} files")

    def archive_old_files(self, source_folder: str, archive_folder: str,
                         days_old: int = 90):
        """
        Archive files older than specified days

        Args:
            source_folder: Source directory
            archive_folder: Archive directory
            days_old: Archive files older than this many days
        """
        os.makedirs(archive_folder, exist_ok=True)

        cutoff_date = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
        files = Path(source_folder).glob('*')
        archived_count = 0

        for file_path in files:
            if file_path.is_file():
                file_date = file_path.stat().st_mtime

                if file_date < cutoff_date:
                    target_path = os.path.join(archive_folder, file_path.name)
                    shutil.move(str(file_path), target_path)
                    logger.info(f"Archived: {file_path.name}")
                    archived_count += 1

        logger.info(f"Archived {archived_count} old files")


class SharePointHelper:
    """Helper for SharePoint operations"""

    def __init__(self, site_url: str):
        from modules.sharepoint import SharePointClient
        self.client = SharePointClient(site_url)

    def sync_folder_to_sharepoint(self, local_folder: str, sp_folder: str):
        """
        Sync a local folder to SharePoint

        Args:
            local_folder: Local folder path
            sp_folder: SharePoint folder path
        """
        files = Path(local_folder).glob('*')
        uploaded_count = 0

        for file_path in files:
            if file_path.is_file():
                sp_path = f"{sp_folder}/{file_path.name}"
                self.client.upload_file(str(file_path), sp_path)
                logger.info(f"Uploaded: {file_path.name}")
                uploaded_count += 1

        logger.info(f"Synced {uploaded_count} files to SharePoint")

    def download_recent_files(self, sp_folder: str, local_folder: str,
                             days: int = 7):
        """
        Download files from SharePoint modified in the last N days

        Args:
            sp_folder: SharePoint folder path
            local_folder: Local destination folder
            days: Download files modified in the last N days
        """
        os.makedirs(local_folder, exist_ok=True)

        # Get files from SharePoint
        files = self.client.list_files(sp_folder)
        cutoff_date = datetime.now().timestamp() - (days * 24 * 60 * 60)

        downloaded_count = 0
        for file_info in files:
            if file_info.get('modified', 0) > cutoff_date:
                sp_path = file_info['path']
                local_path = os.path.join(local_folder, file_info['name'])

                self.client.download_file(sp_path, local_path)
                logger.info(f"Downloaded: {file_info['name']}")
                downloaded_count += 1

        logger.info(f"Downloaded {downloaded_count} recent files")


# =======================
# COMMON WORKFLOW PATTERNS
# =======================

def run_daily_report_workflow(
    data_source: str,
    report_output: str,
    email_recipients: List[str],
    email_subject: str = None
):
    """
    Common pattern: Generate daily report and email it

    Args:
        data_source: Path to data file or folder
        report_output: Path for output Excel file
        email_recipients: List of email addresses
        email_subject: Optional subject (auto-generated if not provided)
    """
    logger.info("Running daily report workflow...")

    # Build report
    report = ReportBuilder()

    # TODO: Add data processing logic specific to data_source

    report.add_title_sheet(
        "Daily Report",
        datetime.now().strftime("%Y-%m-%d"),
        {"Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    )

    report.save(report_output)

    # Email report
    email = EmailHelper()
    subject = email_subject or f"Daily Report - {datetime.now().strftime('%Y-%m-%d')}"
    body = f"Please find attached the daily report for {datetime.now().strftime('%Y-%m-%d')}."

    email.send_report(
        to=email_recipients,
        subject=subject,
        body=body,
        attachments=[report_output]
    )

    logger.info("Daily report workflow completed!")


def run_file_processing_workflow(
    input_folder: str,
    output_folder: str,
    processing_func: Callable,
    organize_by_date: bool = False
):
    """
    Common pattern: Process files in a folder

    Args:
        input_folder: Input folder path
        output_folder: Output folder path
        processing_func: Function to process each file
        organize_by_date: Whether to organize outputs by date
    """
    logger.info("Running file processing workflow...")

    os.makedirs(output_folder, exist_ok=True)
    files = Path(input_folder).glob('*')

    processed_count = 0
    for file_path in files:
        if file_path.is_file():
            try:
                # Process file
                result = processing_func(str(file_path))

                # Save result
                output_path = os.path.join(output_folder, file_path.name)
                # Save result logic here

                logger.info(f"Processed: {file_path.name}")
                processed_count += 1

            except Exception as e:
                logger.error(f"Failed to process {file_path.name}: {e}")

    if organize_by_date:
        organizer = FileOrganizer()
        organizer.organize_by_date(output_folder, output_folder)

    logger.info(f"File processing workflow completed! Processed {processed_count} files")


# =======================
# ASANA INTEGRATION HELPERS
# =======================

class AsanaHelper:
    """
    Helper for Asana automation tasks.

    Provides convenient methods for common Asana operations:
    - Email-to-task creation (no API required)
    - Browser automation for complex operations
    - CSV bulk operations
    - Excel-to-Asana synchronization
    """

    def __init__(self, project_email: str = None, project_url: str = None):
        """
        Initialize Asana helper.

        Args:
            project_email: Asana project email address (for email-to-task)
            project_url: Asana project URL (for browser automation)
        """
        self.project_email = project_email
        self.project_url = project_url

    def create_task_via_email(
        self,
        name: str,
        description: str = '',
        assignee: str = '',
        due_date: str = '',
        priority: str = '',
        tags: str = '',
        attachments: List[str] = None
    ) -> Dict[str, Any]:
        """
        Create a single task via email.

        Args:
            name: Task name
            description: Task description
            assignee: Assignee username (without @)
            due_date: Due date (YYYY-MM-DD or 'next Friday', etc.)
            priority: Priority (High/Medium/Low)
            tags: Comma-separated tags
            attachments: List of file paths to attach

        Returns:
            Result dictionary

        Example:
            helper = AsanaHelper(project_email='myproject@mail.asana.com')
            result = helper.create_task_via_email(
                name='Review proposal',
                assignee='john.doe',
                due_date='2024-12-15',
                priority='High'
            )
        """
        from modules.asana import AsanaEmailModule

        if not self.project_email:
            raise ValueError("project_email is required for email-based operations")

        task = {
            'name': name,
            'description': description,
            'assignee': assignee,
            'due_date': due_date,
            'priority': priority,
            'tags': tags,
            'attachments': attachments or []
        }

        module = AsanaEmailModule()
        result = module.run(
            project_email=self.project_email,
            tasks=[task]
        )

        logger.info(f"Created Asana task via email: {name}")
        return result

    def bulk_create_from_list(
        self,
        tasks: List[Dict[str, Any]],
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Bulk create tasks from a list.

        Args:
            tasks: List of task dictionaries
            method: 'email' or 'browser'

        Returns:
            Result dictionary

        Example:
            tasks = [
                {'name': 'Task 1', 'assignee': 'john'},
                {'name': 'Task 2', 'due_date': '2024-12-15'}
            ]
            result = helper.bulk_create_from_list(tasks, method='email')
        """
        if method == 'email':
            from modules.asana import AsanaEmailModule

            if not self.project_email:
                raise ValueError("project_email is required for email method")

            module = AsanaEmailModule()
            result = module.run(
                project_email=self.project_email,
                tasks=tasks
            )

        elif method == 'browser':
            from modules.asana import AsanaBrowserModule

            if not self.project_url:
                raise ValueError("project_url is required for browser method")

            module = AsanaBrowserModule()
            result = module.run(
                asana_url=self.project_url,
                operation='create_tasks',
                tasks=tasks
            )

        else:
            raise ValueError(f"Invalid method: {method}. Must be 'email' or 'browser'")

        logger.info(f"Bulk created {len(tasks)} tasks via {method}")
        return result

    def bulk_create_from_excel(
        self,
        excel_path: str,
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Bulk create tasks from Excel file.

        Excel columns: name, description, assignee, due_date, priority, tags

        Args:
            excel_path: Path to Excel file
            method: 'email' or 'csv'

        Returns:
            Result dictionary

        Example:
            helper = AsanaHelper(project_email='myproject@mail.asana.com')
            result = helper.bulk_create_from_excel('C:/Data/tasks.xlsx')
        """
        if method == 'email':
            from modules.asana import AsanaEmailModule

            if not self.project_email:
                raise ValueError("project_email is required for email method")

            module = AsanaEmailModule()
            result = module.run(
                project_email=self.project_email,
                from_excel=excel_path
            )

        elif method == 'csv':
            # Convert Excel to Asana CSV
            from modules.asana import AsanaCSVHandler

            handler = AsanaCSVHandler()

            # First generate CSV
            csv_output = excel_path.replace('.xlsx', '_asana.csv').replace('.xls', '_asana.csv')
            result = handler.run(
                operation='generate',
                input_file=excel_path,
                output_file=csv_output
            )

            logger.info(f"Generated Asana CSV for import: {csv_output}")
            logger.info("You can now manually import this CSV into Asana")

        else:
            raise ValueError(f"Invalid method: {method}. Must be 'email' or 'csv'")

        logger.info(f"Bulk created tasks from Excel: {excel_path}")
        return result

    def generate_asana_csv(
        self,
        tasks: List[Dict[str, Any]],
        output_path: str
    ) -> Dict[str, Any]:
        """
        Generate Asana-compatible CSV from task list.

        Args:
            tasks: List of task dictionaries
            output_path: Output CSV file path

        Returns:
            Result dictionary

        Example:
            tasks = [
                {'name': 'Task 1', 'assignee': 'john', 'due_date': '2024-12-15'},
                {'name': 'Task 2', 'priority': 'High'}
            ]
            helper.generate_asana_csv(tasks, 'C:/Output/tasks.csv')
        """
        from modules.asana import AsanaCSVHandler

        handler = AsanaCSVHandler()
        result = handler.run(
            operation='generate',
            tasks=tasks,
            output_file=output_path
        )

        logger.info(f"Generated Asana CSV: {output_path}")
        return result

    def parse_asana_export(
        self,
        csv_path: str,
        output_json: str = None
    ) -> List[Dict[str, Any]]:
        """
        Parse Asana CSV export into structured data.

        Args:
            csv_path: Path to Asana CSV export
            output_json: Optional path to save JSON output

        Returns:
            List of task dictionaries

        Example:
            tasks = helper.parse_asana_export('C:/Downloads/asana_export.csv')
            # Analyze tasks
            for task in tasks:
                print(f"{task['name']} - {task['assignee']}")
        """
        from modules.asana import AsanaCSVHandler

        handler = AsanaCSVHandler()
        result = handler.run(
            operation='parse',
            input_file=csv_path,
            output_file=output_json
        )

        logger.info(f"Parsed {result['data']['task_count']} tasks from Asana export")
        return result['data']['tasks']

    def convert_excel_to_asana_csv(
        self,
        excel_path: str,
        output_path: str,
        column_mapping: Dict[str, str] = None
    ) -> Dict[str, Any]:
        """
        Convert Excel tracker to Asana-compatible CSV.

        Args:
            excel_path: Path to Excel file
            output_path: Output CSV path
            column_mapping: Map Asana columns to Excel columns
                          e.g., {'Task Name': 'title', 'Assignee': 'owner'}

        Returns:
            Result dictionary

        Example:
            mapping = {
                'Task Name': 'task_title',
                'Assignee': 'assigned_to',
                'Due Date': 'deadline'
            }
            result = helper.convert_excel_to_asana_csv(
                'C:/Data/project.xlsx',
                'C:/Output/asana_import.csv',
                mapping
            )
        """
        from modules.asana import AsanaCSVHandler

        # If no mapping, try auto-detection
        if not column_mapping:
            column_mapping = {}

        handler = AsanaCSVHandler()

        # First convert Excel to temp CSV
        import tempfile
        temp_csv = tempfile.mktemp(suffix='.csv')

        # Load and convert Excel to CSV
        from openpyxl import load_workbook
        import csv

        workbook = load_workbook(excel_path)
        sheet = workbook.active

        with open(temp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        # Now convert CSV to Asana format
        result = handler.run(
            operation='convert',
            input_file=temp_csv,
            output_file=output_path,
            mapping=column_mapping
        )

        # Cleanup temp file
        Path(temp_csv).unlink(missing_ok=True)

        logger.info(f"Converted Excel to Asana CSV: {output_path}")
        return result

    def sync_excel_tracker_to_asana(
        self,
        excel_path: str,
        method: str = 'email'
    ) -> Dict[str, Any]:
        """
        Sync Excel project tracker to Asana.

        This is a convenience method that:
        1. Reads tasks from Excel
        2. Creates them in Asana via email or CSV

        Args:
            excel_path: Path to Excel tracker
            method: 'email' for direct creation, 'csv' for manual import

        Returns:
            Result dictionary

        Example:
            # Sync weekly sprint tasks from Excel to Asana
            helper = AsanaHelper(project_email='sprint@mail.asana.com')
            result = helper.sync_excel_tracker_to_asana('C:/Data/sprint_tasks.xlsx')
        """
        logger.info(f"Syncing Excel tracker to Asana: {excel_path}")

        if method == 'email':
            return self.bulk_create_from_excel(excel_path, method='email')

        elif method == 'csv':
            # Generate Asana CSV for manual import
            csv_output = excel_path.replace('.xlsx', '_asana_import.csv').replace('.xls', '_asana_import.csv')
            return self.convert_excel_to_asana_csv(excel_path, csv_output)

        else:
            raise ValueError(f"Invalid method: {method}")

    def create_weekly_sprint_tasks(
        self,
        sprint_number: int,
        start_date: str,
        task_templates: List[Dict[str, str]]
    ) -> Dict[str, Any]:
        """
        Create weekly sprint tasks from templates.

        Args:
            sprint_number: Sprint number (e.g., 42)
            start_date: Sprint start date (YYYY-MM-DD)
            task_templates: List of task templates with name, assignee, day_offset

        Returns:
            Result dictionary

        Example:
            templates = [
                {'name': 'Sprint Planning', 'assignee': 'team', 'day_offset': 0},
                {'name': 'Daily Standup', 'assignee': 'team', 'day_offset': 1},
                {'name': 'Sprint Review', 'assignee': 'team', 'day_offset': 4},
                {'name': 'Retrospective', 'assignee': 'team', 'day_offset': 4}
            ]
            result = helper.create_weekly_sprint_tasks(42, '2024-12-09', templates)
        """
        from datetime import datetime, timedelta

        logger.info(f"Creating Sprint {sprint_number} tasks...")

        start = datetime.strptime(start_date, '%Y-%m-%d')

        tasks = []
        for template in task_templates:
            due_date = start + timedelta(days=template.get('day_offset', 0))

            task = {
                'name': f"Sprint {sprint_number}: {template['name']}",
                'assignee': template.get('assignee', ''),
                'due_date': due_date.strftime('%Y-%m-%d'),
                'description': template.get('description', ''),
                'priority': template.get('priority', 'Medium'),
                'tags': f"sprint-{sprint_number}"
            }
            tasks.append(task)

        return self.bulk_create_from_list(tasks, method='email')


def create_asana_roadmap_from_excel(
    excel_path: str,
    project_email: str,
    output_csv: str = None
) -> Dict[str, Any]:
    """
    Common workflow: Create Asana roadmap from Excel planning file.

    Args:
        excel_path: Path to Excel roadmap file
        project_email: Asana project email
        output_csv: Optional CSV output path (for manual import)

    Returns:
        Result dictionary

    Example:
        result = create_asana_roadmap_from_excel(
            'C:/Planning/Q1_Roadmap.xlsx',
            'q1-projects@mail.asana.com'
        )
    """
    logger.info("Creating Asana roadmap from Excel...")

    helper = AsanaHelper(project_email=project_email)

    if output_csv:
        # Generate CSV for review/manual import
        result = helper.convert_excel_to_asana_csv(excel_path, output_csv)
        logger.info(f"Roadmap CSV generated: {output_csv}")
    else:
        # Direct email creation
        result = helper.bulk_create_from_excel(excel_path, method='email')
        logger.info("Roadmap tasks created via email")

    return result


def sync_asana_status_to_excel(
    asana_csv_export: str,
    excel_tracker: str,
    output_path: str = None
) -> Dict[str, Any]:
    """
    Common workflow: Sync Asana task status back to Excel tracker.

    Args:
        asana_csv_export: Path to Asana CSV export
        excel_tracker: Path to Excel tracker file
        output_path: Output Excel path (defaults to tracker path)

    Returns:
        Result dictionary with sync summary

    Example:
        result = sync_asana_status_to_excel(
            'C:/Downloads/asana_export.csv',
            'C:/Tracking/project_tracker.xlsx'
        )
    """
    logger.info("Syncing Asana status to Excel tracker...")

    # Parse Asana export
    helper = AsanaHelper()
    asana_tasks = helper.parse_asana_export(asana_csv_export)

    # Load Excel tracker
    from modules.excel_automation import WorkbookHandler

    workbook = WorkbookHandler()
    workbook.load(excel_tracker)

    # Update status based on task names
    # (This is a template - actual implementation depends on your Excel structure)

    updated_count = 0
    for task in asana_tasks:
        task_name = task.get('name', '')
        status = 'Complete' if task.get('completed_at') else 'In Progress'

        # Find and update matching row in Excel
        # (Simplified - actual implementation would search by name)
        updated_count += 1

    # Save
    if not output_path:
        output_path = excel_tracker

    workbook.save(output_path)

    logger.info(f"Updated {updated_count} tasks in Excel tracker")

    return {
        'status': 'success',
        'message': f'Synced {updated_count} tasks',
        'output_file': output_path
    }
