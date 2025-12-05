"""
Asana Email Integration Module

Creates Asana tasks by sending emails to project-specific addresses.
This method requires no API keys and works within corporate restrictions.

How it works:
- Each Asana project has a unique email address (found in Project Settings)
- Email subject becomes task name
- Email body becomes task description
- Special syntax for assignee (@username) and due dates
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from modules.base_module import BaseModule, ModuleStatus

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False


class AsanaEmailModule(BaseModule):
    """
    Create Asana tasks via email-to-task integration.

    Configuration:
        - project_email: Asana project email address
        - tasks: List of task dictionaries to create
        - from_excel: Optional Excel file path with tasks

    Task dictionary format:
        {
            'name': 'Task name',
            'description': 'Task description (optional)',
            'assignee': 'username' (optional, will add @username to body),
            'due_date': 'YYYY-MM-DD' (optional),
            'priority': 'High/Medium/Low' (optional, adds ! or !! to name)
        }
    """

    def __init__(self):
        super().__init__(
            name="AsanaEmailModule",
            description="Create Asana tasks via email integration",
            version="1.0.0",
            category="asana"
        )

        self.outlook = None
        self.project_email = None
        self.tasks = []

    def configure(self, **kwargs) -> bool:
        """
        Configure the module.

        Args:
            project_email: Asana project email address
            tasks: List of task dictionaries
            from_excel: Optional Excel file path with tasks
        """
        try:
            # Validate win32com availability
            if not WIN32COM_AVAILABLE:
                self.log_error("pywin32 is required but not installed")
                return False

            # Get project email
            self.project_email = kwargs.get('project_email')
            if not self.project_email:
                self.log_error("project_email is required")
                return False

            # Get tasks from kwargs or Excel
            if 'from_excel' in kwargs:
                self.tasks = self._load_tasks_from_excel(kwargs['from_excel'])
            elif 'tasks' in kwargs:
                self.tasks = kwargs['tasks']
            else:
                self.log_error("Either 'tasks' or 'from_excel' must be provided")
                return False

            # Store config
            self.set_config('project_email', self.project_email)
            self.set_config('task_count', len(self.tasks))

            self.log_info(f"Configured with {len(self.tasks)} tasks for project: {self.project_email}")
            return True

        except Exception as e:
            self.log_error(f"Configuration failed: {e}")
            return False

    def validate(self) -> bool:
        """Validate configuration."""
        try:
            # Validate project email format
            if '@' not in self.project_email or 'asana' not in self.project_email.lower():
                self.log_warning(f"Email doesn't look like an Asana project email: {self.project_email}")

            # Validate tasks
            if not self.tasks:
                self.log_error("No tasks to create")
                return False

            # Validate each task
            for i, task in enumerate(self.tasks):
                if not isinstance(task, dict):
                    self.log_error(f"Task {i} is not a dictionary")
                    return False

                if 'name' not in task:
                    self.log_error(f"Task {i} missing required 'name' field")
                    return False

            self.log_info("Validation passed")
            return True

        except Exception as e:
            self.log_error(f"Validation failed: {e}")
            return False

    def execute(self) -> Dict[str, Any]:
        """Create Asana tasks by sending emails."""
        try:
            # Connect to Outlook
            self._connect_outlook()

            created_tasks = []
            failed_tasks = []

            # Create each task
            for i, task in enumerate(self.tasks, 1):
                self.log_info(f"Creating task {i}/{len(self.tasks)}: {task['name']}")

                try:
                    self._send_task_email(task)
                    created_tasks.append(task['name'])
                    self.log_info(f"✓ Created: {task['name']}")

                except Exception as e:
                    self.log_error(f"✗ Failed to create task '{task['name']}': {e}")
                    failed_tasks.append({
                        'name': task['name'],
                        'error': str(e)
                    })

            # Build result
            result_data = {
                'created_count': len(created_tasks),
                'failed_count': len(failed_tasks),
                'created_tasks': created_tasks,
                'failed_tasks': failed_tasks,
                'project_email': self.project_email
            }

            if failed_tasks:
                return self.success_result(
                    result_data,
                    f"Created {len(created_tasks)} tasks, {len(failed_tasks)} failed"
                )
            else:
                return self.success_result(
                    result_data,
                    f"Successfully created all {len(created_tasks)} tasks"
                )

        except Exception as e:
            self.log_error(f"Execution failed: {e}")
            return self.error_result(f"Failed to create tasks: {e}")

    def _connect_outlook(self):
        """Connect to Outlook."""
        try:
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            self.log_info("Connected to Outlook")
        except Exception as e:
            self.log_error(f"Failed to connect to Outlook: {e}")
            raise

    def _send_task_email(self, task: Dict[str, Any]):
        """
        Send email to create a task in Asana.

        Args:
            task: Task dictionary with name, description, assignee, due_date
        """
        # Build email subject (task name)
        subject = task['name']

        # Add priority indicator
        priority = task.get('priority', '').lower()
        if priority == 'high':
            subject = '!! ' + subject
        elif priority == 'medium':
            subject = '! ' + subject

        # Build email body (task description)
        body_parts = []

        # Add description if provided
        if task.get('description'):
            body_parts.append(task['description'])
            body_parts.append('')  # Blank line

        # Add assignee if provided
        if task.get('assignee'):
            assignee = task['assignee']
            # Add @ if not already present
            if not assignee.startswith('@'):
                assignee = '@' + assignee
            body_parts.append(f"Assignee: {assignee}")

        # Add due date if provided
        if task.get('due_date'):
            due_date = task['due_date']
            # Format: "Due: YYYY-MM-DD" or "Due: next Friday"
            body_parts.append(f"Due: {due_date}")

        # Add tags if provided
        if task.get('tags'):
            tags = task['tags']
            if isinstance(tags, list):
                tags = ', '.join(tags)
            body_parts.append(f"Tags: {tags}")

        # Add notes if provided
        if task.get('notes'):
            body_parts.append('')
            body_parts.append('Notes:')
            body_parts.append(task['notes'])

        body = '\n'.join(body_parts)

        # Create and send email
        mail = self.outlook.CreateItem(0)  # 0 = MailItem
        mail.To = self.project_email
        mail.Subject = subject
        mail.Body = body

        # Add attachments if provided
        if task.get('attachments'):
            attachments = task['attachments']
            if not isinstance(attachments, list):
                attachments = [attachments]

            for attachment in attachments:
                try:
                    mail.Attachments.Add(attachment)
                    self.log_debug(f"Added attachment: {attachment}")
                except Exception as e:
                    self.log_warning(f"Failed to add attachment {attachment}: {e}")

        # Send email
        mail.Send()
        self.log_debug(f"Sent email to {self.project_email}")

    def _load_tasks_from_excel(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        Load tasks from an Excel file.

        Expected columns: name, description, assignee, due_date, priority, tags

        Args:
            excel_path: Path to Excel file

        Returns:
            List of task dictionaries
        """
        try:
            from openpyxl import load_workbook

            self.log_info(f"Loading tasks from Excel: {excel_path}")

            workbook = load_workbook(excel_path)
            sheet = workbook.active

            # Read header row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value.lower() if cell.value else '')

            # Validate required columns
            if 'name' not in headers:
                raise ValueError("Excel must have 'name' column")

            # Read tasks
            tasks = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(row):
                    continue

                # Build task dictionary
                task = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        task[headers[i]] = str(value)

                if task.get('name'):
                    tasks.append(task)

            self.log_info(f"Loaded {len(tasks)} tasks from Excel")
            return tasks

        except Exception as e:
            self.log_error(f"Failed to load tasks from Excel: {e}")
            raise

    def create_single_task(self, project_email: str, name: str,
                          description: str = '', assignee: str = '',
                          due_date: str = '', priority: str = '') -> Dict[str, Any]:
        """
        Convenience method to create a single task.

        Args:
            project_email: Asana project email
            name: Task name
            description: Task description
            assignee: Assignee username
            due_date: Due date (YYYY-MM-DD format)
            priority: Priority (High/Medium/Low)

        Returns:
            Result dictionary
        """
        task = {
            'name': name,
            'description': description,
            'assignee': assignee,
            'due_date': due_date,
            'priority': priority
        }

        return self.run(
            project_email=project_email,
            tasks=[task]
        )

    def bulk_create_from_excel(self, project_email: str, excel_path: str) -> Dict[str, Any]:
        """
        Convenience method to bulk create tasks from Excel.

        Args:
            project_email: Asana project email
            excel_path: Path to Excel file with tasks

        Returns:
            Result dictionary
        """
        return self.run(
            project_email=project_email,
            from_excel=excel_path
        )


# Example usage
if __name__ == '__main__':
    # Example 1: Create single task
    module = AsanaEmailModule()
    result = module.create_single_task(
        project_email='your-project@mail.asana.com',
        name='Review Q4 Budget',
        description='Review and approve the Q4 budget proposal',
        assignee='john.doe',
        due_date='2024-12-15',
        priority='High'
    )
    print(result)

    # Example 2: Create multiple tasks
    tasks = [
        {
            'name': 'Update documentation',
            'description': 'Update API documentation for new endpoints',
            'assignee': 'jane.smith',
            'due_date': '2024-12-10'
        },
        {
            'name': 'Code review PR #123',
            'priority': 'High'
        }
    ]

    module2 = AsanaEmailModule()
    result2 = module2.run(
        project_email='your-project@mail.asana.com',
        tasks=tasks
    )
    print(result2)
