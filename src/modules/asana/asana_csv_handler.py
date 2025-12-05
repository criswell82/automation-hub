"""
Asana CSV Handler

Generate CSV files compatible with Asana's CSV import feature.
Also processes CSV exports from Asana for analysis and reporting.

Asana CSV Format:
- Required columns: Task Name (or Name)
- Optional columns: Assignee, Due Date, Notes, Priority, Section, Tags, Projects, etc.

Use cases:
- Bulk task creation from Excel/CSV data
- Generate project plans for import
- Process Asana CSV exports for reporting
- Sync Excel trackers with Asana via CSV
"""

import csv
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import sys

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from modules.base_module import BaseModule


class AsanaCSVHandler(BaseModule):
    """
    Handle Asana CSV import/export operations.

    Configuration:
        - operation: 'generate' | 'parse' | 'convert'
        - input_file: Source file path (for parse/convert)
        - output_file: Destination file path
        - tasks: Task data (for generate operation)
        - mapping: Column mapping (for convert operation)
    """

    def __init__(self) -> None:
        super().__init__(
            name="AsanaCSVHandler",
            description="Generate and process Asana CSV files",
            version="1.0.0",
            category="asana"
        )

        self.operation = None
        self.input_file = None
        self.output_file = None
        self.tasks = []
        self.mapping = {}

        # Asana CSV column names (standard format)
        self.asana_columns = [
            'Task Name',
            'Assignee',
            'Due Date',
            'Notes',
            'Priority',
            'Section',
            'Tags',
            'Projects',
            'Created At',
            'Completed At',
            'Custom Field 1',
            'Custom Field 2'
        ]

    def configure(self, **kwargs) -> bool:
        """
        Configure the module.

        Args:
            operation: 'generate', 'parse', or 'convert'
            input_file: Source file (for parse/convert)
            output_file: Destination file
            tasks: Task data (for generate)
            mapping: Column mapping (for convert)
        """
        try:
            self.operation = kwargs.get('operation', 'generate')
            self.input_file = kwargs.get('input_file')
            self.output_file = kwargs.get('output_file')
            self.tasks = kwargs.get('tasks', [])
            self.mapping = kwargs.get('mapping', {})

            # Store config
            self.set_config('operation', self.operation)

            self.log_info(f"Configured for operation: {self.operation}")
            return True

        except Exception as e:
            self.log_error(f"Configuration failed: {e}")
            return False

    def validate(self) -> bool:
        """Validate configuration."""
        try:
            valid_operations = ['generate', 'parse', 'convert']
            if self.operation not in valid_operations:
                self.log_error(f"Invalid operation: {self.operation}")
                return False

            # Validate based on operation
            if self.operation == 'generate':
                if not self.tasks and not self.input_file:
                    self.log_error("Generate operation requires 'tasks' or 'input_file'")
                    return False
                if not self.output_file:
                    self.log_error("Generate operation requires 'output_file'")
                    return False

            elif self.operation in ['parse', 'convert']:
                if not self.input_file:
                    self.log_error(f"{self.operation} operation requires 'input_file'")
                    return False

            self.log_info("Validation passed")
            return True

        except Exception as e:
            self.log_error(f"Validation failed: {e}")
            return False

    def execute(self) -> Dict[str, Any]:
        """Execute the configured operation."""
        try:
            if self.operation == 'generate':
                return self._generate_csv()
            elif self.operation == 'parse':
                return self._parse_csv()
            elif self.operation == 'convert':
                return self._convert_csv()
            else:
                raise ValueError(f"Unknown operation: {self.operation}")

        except Exception as e:
            self.log_error(f"Execution failed: {e}")
            return self.error_result(f"Operation failed: {e}")

    def _generate_csv(self) -> Dict[str, Any]:
        """
        Generate Asana-compatible CSV from task data.

        Returns:
            Result dictionary
        """
        try:
            # Load tasks from input file if provided
            if self.input_file and not self.tasks:
                self.tasks = self._load_tasks_from_file(self.input_file)

            self.log_info(f"Generating Asana CSV with {len(self.tasks)} tasks...")

            # Prepare CSV data
            csv_data = []
            for task in self.tasks:
                csv_row = self._task_to_csv_row(task)
                csv_data.append(csv_row)

            # Write CSV file
            self._write_csv(self.output_file, csv_data)

            return self.success_result(
                {
                    'output_file': self.output_file,
                    'task_count': len(csv_data)
                },
                f"Generated CSV with {len(csv_data)} tasks"
            )

        except Exception as e:
            self.log_error(f"Failed to generate CSV: {e}")
            raise

    def _parse_csv(self) -> Dict[str, Any]:
        """
        Parse Asana CSV export into structured data.

        Returns:
            Result dictionary with parsed tasks
        """
        try:
            self.log_info(f"Parsing Asana CSV: {self.input_file}")

            tasks = []
            with open(self.input_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)

                for row in reader:
                    task = self._csv_row_to_task(row)
                    tasks.append(task)

            self.log_info(f"Parsed {len(tasks)} tasks from CSV")

            # Save to output file if specified
            if self.output_file:
                self._save_tasks_to_file(tasks, self.output_file)

            return self.success_result(
                {
                    'tasks': tasks,
                    'task_count': len(tasks),
                    'output_file': self.output_file
                },
                f"Parsed {len(tasks)} tasks"
            )

        except Exception as e:
            self.log_error(f"Failed to parse CSV: {e}")
            raise

    def _convert_csv(self) -> Dict[str, Any]:
        """
        Convert custom CSV format to Asana format using column mapping.

        Example mapping:
            {
                'Task Name': 'task_title',
                'Assignee': 'assigned_to',
                'Due Date': 'deadline'
            }

        Returns:
            Result dictionary
        """
        try:
            self.log_info(f"Converting CSV with mapping: {self.mapping}")

            # Read input CSV
            tasks = []
            with open(self.input_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)

                for row in reader:
                    # Map columns to Asana format
                    task = {}
                    for asana_col, source_col in self.mapping.items():
                        if source_col in row:
                            task[asana_col] = row[source_col]

                    # Add any unmapped columns
                    for key, value in row.items():
                        if key not in self.mapping.values() and key not in task:
                            # Try to intelligently map
                            if 'name' in key.lower() or 'title' in key.lower():
                                task['Task Name'] = value
                            elif 'assign' in key.lower():
                                task['Assignee'] = value
                            elif 'date' in key.lower() or 'due' in key.lower():
                                task['Due Date'] = value
                            elif 'note' in key.lower() or 'desc' in key.lower():
                                task['Notes'] = value

                    tasks.append(task)

            # Write Asana-format CSV
            if not self.output_file:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.output_file = str(self.get_output_path(f"asana_import_{timestamp}.csv"))

            csv_data = [task for task in tasks]
            self._write_csv(self.output_file, csv_data)

            self.log_info(f"Converted {len(tasks)} tasks to Asana format")

            return self.success_result(
                {
                    'output_file': self.output_file,
                    'task_count': len(tasks)
                },
                f"Converted {len(tasks)} tasks"
            )

        except Exception as e:
            self.log_error(f"Failed to convert CSV: {e}")
            raise

    def _task_to_csv_row(self, task: Dict[str, Any]) -> Dict[str, str]:
        """
        Convert task dictionary to Asana CSV row.

        Args:
            task: Task dictionary

        Returns:
            CSV row dictionary with Asana column names
        """
        row = {}

        # Map common fields
        row['Task Name'] = task.get('name', task.get('task_name', ''))
        row['Assignee'] = task.get('assignee', task.get('assigned_to', ''))
        row['Due Date'] = task.get('due_date', task.get('deadline', ''))
        row['Notes'] = task.get('notes', task.get('description', ''))
        row['Priority'] = task.get('priority', '')
        row['Section'] = task.get('section', task.get('category', ''))
        row['Tags'] = task.get('tags', '')
        row['Projects'] = task.get('projects', '')

        # Handle tags if list
        if isinstance(row['Tags'], list):
            row['Tags'] = ', '.join(row['Tags'])

        return row

    def _csv_row_to_task(self, row: Dict[str, str]) -> Dict[str, Any]:
        """
        Convert Asana CSV row to task dictionary.

        Args:
            row: CSV row dictionary

        Returns:
            Task dictionary
        """
        task = {
            'name': row.get('Task Name', row.get('Name', '')),
            'assignee': row.get('Assignee', ''),
            'due_date': row.get('Due Date', ''),
            'notes': row.get('Notes', ''),
            'priority': row.get('Priority', ''),
            'section': row.get('Section', ''),
            'tags': row.get('Tags', ''),
            'projects': row.get('Projects', ''),
            'created_at': row.get('Created At', ''),
            'completed_at': row.get('Completed At', '')
        }

        # Convert tags to list if present
        if task['tags']:
            task['tags'] = [tag.strip() for tag in task['tags'].split(',')]

        return task

    def _write_csv(self, filepath: str, data: List[Dict[str, Any]]) -> None:
        """
        Write CSV file in Asana format.

        Args:
            filepath: Output file path
            data: List of row dictionaries
        """
        if not data:
            self.log_warning("No data to write to CSV")
            return

        # Ensure directory exists
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)

        # Determine columns from first row + standard columns
        columns = list(self.asana_columns)
        for row in data:
            for key in row.keys():
                if key not in columns:
                    columns.append(key)

        # Write CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)

        self.log_info(f"CSV file written: {filepath}")

    def _load_tasks_from_file(self, filepath: str) -> List[Dict[str, Any]]:
        """
        Load tasks from Excel or JSON file.

        Args:
            filepath: Input file path

        Returns:
            List of task dictionaries
        """
        ext = Path(filepath).suffix.lower()

        if ext in ['.xlsx', '.xls']:
            return self._load_from_excel(filepath)
        elif ext == '.json':
            return self._load_from_json(filepath)
        elif ext == '.csv':
            return self._load_from_csv(filepath)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    def _load_from_excel(self, filepath: str) -> List[Dict[str, Any]]:
        """Load tasks from Excel file."""
        from openpyxl import load_workbook

        workbook = load_workbook(filepath)
        sheet = workbook.active

        # Read header
        headers = [cell.value for cell in sheet[1]]

        # Read tasks
        tasks = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            task = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    task[headers[i]] = value
            if task:
                tasks.append(task)

        return tasks

    def _load_from_json(self, filepath: str) -> List[Dict[str, Any]]:
        """Load tasks from JSON file."""
        import json

        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if isinstance(data, list):
            return data
        elif isinstance(data, dict) and 'tasks' in data:
            return data['tasks']
        else:
            raise ValueError("JSON must be a list of tasks or dict with 'tasks' key")

    def _load_from_csv(self, filepath: str) -> List[Dict[str, Any]]:
        """Load tasks from CSV file."""
        tasks = []
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                tasks.append(dict(row))
        return tasks

    def _save_tasks_to_file(self, tasks: List[Dict[str, Any]], filepath: str) -> None:
        """Save tasks to JSON file."""
        import json

        Path(filepath).parent.mkdir(parents=True, exist_ok=True)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(tasks, f, indent=2)

        self.log_info(f"Tasks saved to: {filepath}")


# Example usage
if __name__ == '__main__':
    # Example 1: Generate Asana CSV from task list
    tasks = [
        {
            'name': 'Design Homepage Mockup',
            'assignee': 'jane.smith',
            'due_date': '2024-12-15',
            'notes': 'Create mobile and desktop versions',
            'priority': 'High',
            'section': 'Design',
            'tags': ['UI', 'Homepage']
        },
        {
            'name': 'Implement User Authentication',
            'assignee': 'john.doe',
            'due_date': '2024-12-20',
            'priority': 'High',
            'section': 'Development'
        }
    ]

    handler = AsanaCSVHandler()
    result = handler.run(
        operation='generate',
        tasks=tasks,
        output_file='C:/Output/asana_import.csv'
    )
    print(result)

    # Example 2: Parse Asana CSV export
    handler2 = AsanaCSVHandler()
    result2 = handler2.run(
        operation='parse',
        input_file='C:/Downloads/asana_export.csv',
        output_file='C:/Output/parsed_tasks.json'
    )
    print(result2)

    # Example 3: Convert custom CSV to Asana format
    mapping = {
        'Task Name': 'title',
        'Assignee': 'owner',
        'Due Date': 'deadline',
        'Notes': 'description'
    }

    handler3 = AsanaCSVHandler()
    result3 = handler3.run(
        operation='convert',
        input_file='C:/Data/project_tasks.csv',
        output_file='C:/Output/asana_import.csv',
        mapping=mapping
    )
    print(result3)
