"""
Workbook Handler for Excel automation.
Handles reading, writing, and manipulating Excel workbooks.
"""

import logging
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class WorkbookHandler:
    """Handles Excel workbook operations using openpyxl."""

    def __init__(self, filepath: Optional[str] = None):
        """
        Initialize workbook handler.

        Args:
            filepath: Path to existing workbook or None for new workbook
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel automation")

        self.logger = logging.getLogger(__name__)
        self.filepath = filepath
        self.workbook: Optional[Workbook] = None

        if filepath and Path(filepath).exists():
            self.load(filepath)
        else:
            self.workbook = Workbook()

    def load(self, filepath: str):
        """Load an existing workbook."""
        try:
            self.workbook = load_workbook(filepath)
            self.filepath = filepath
            self.logger.info(f"Loaded workbook: {filepath}")
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {e}")
            raise

    def save(self, filepath: Optional[str] = None):
        """Save the workbook."""
        save_path = filepath or self.filepath

        if not save_path:
            raise ValueError("No filepath specified for saving")

        try:
            self.workbook.save(save_path)
            self.filepath = save_path
            self.logger.info(f"Saved workbook: {save_path}")
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            raise

    def create_sheet(self, name: str, index: Optional[int] = None):
        """Create a new worksheet."""
        if index is not None:
            self.workbook.create_sheet(name, index)
        else:
            self.workbook.create_sheet(name)
        self.logger.info(f"Created sheet: {name}")

    def get_sheet(self, name: str):
        """Get a worksheet by name."""
        return self.workbook[name]

    def delete_sheet(self, name: str):
        """Delete a worksheet."""
        del self.workbook[name]
        self.logger.info(f"Deleted sheet: {name}")

    def read_data(self, sheet_name: str, range_str: str) -> List[List[Any]]:
        """Read data from a range."""
        sheet = self.get_sheet(sheet_name)
        data = []

        for row in sheet[range_str]:
            data.append([cell.value for cell in row])

        return data

    def write_data(self, sheet_name: str, data: List[List[Any]], start_cell: str = 'A1'):
        """Write data to worksheet."""
        sheet = self.get_sheet(sheet_name)

        # Parse start cell
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_letter, row_num = coordinate_from_string(start_cell)
        start_col = column_index_from_string(col_letter)

        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = sheet.cell(row=row_num + row_idx, column=start_col + col_idx)
                cell.value = value

        self.logger.info(f"Wrote {len(data)} rows to {sheet_name}")

    def format_cells(self, sheet_name: str, range_str: str, **kwargs):
        """Format cells with styling."""
        sheet = self.get_sheet(sheet_name)

        # Extract formatting options
        font_name = kwargs.get('font_name')
        font_size = kwargs.get('font_size')
        bold = kwargs.get('bold', False)
        italic = kwargs.get('italic', False)
        fill_color = kwargs.get('fill_color')
        border = kwargs.get('border', False)

        for row in sheet[range_str]:
            for cell in row:
                if font_name or font_size or bold or italic:
                    cell.font = Font(
                        name=font_name or 'Calibri',
                        size=font_size or 11,
                        bold=bold,
                        italic=italic
                    )

                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

                if border:
                    thin = Side(border_style='thin', color='000000')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_formula(self, sheet_name: str, cell: str, formula: str):
        """Add formula to a cell."""
        sheet = self.get_sheet(sheet_name)
        sheet[cell] = formula
        self.logger.info(f"Added formula to {cell}: {formula}")

    def auto_size_columns(self, sheet_name: str):
        """Auto-size columns based on content."""
        sheet = self.get_sheet(sheet_name)

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        return self.workbook.sheetnames

    def close(self):
        """Close the workbook."""
        self.workbook = None
        self.logger.info("Workbook closed")
