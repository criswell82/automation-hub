"""
WORKFLOW_META:
  name: Quarterly Roadmap Import
  description: Import quarterly roadmap from Excel into Asana with proper sections and dates
  category: Asana
  version: 1.0.0
  author: Automation Hub
  parameters:
    project_email:
      type: string
      description: Asana project email address
      required: true
      default: "q1-roadmap@mail.asana.com"
    roadmap_file:
      type: string
      description: Path to Excel roadmap file
      required: true
      default: "C:/Planning/Q1_Roadmap.xlsx"
    quarter:
      type: string
      description: Quarter (e.g., Q1 2024)
      required: true
      default: "Q1 2025"
"""

import sys
from pathlib import Path
from datetime import datetime

# Add src directory to path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from src.modules.base_module import BaseModule
from src.core.logging_config import get_logger
from src.utils.workflow_helpers import AsanaHelper

logger = get_logger(__name__)


class QuarterlyRoadmapWorkflow(BaseModule):
    """Import quarterly roadmap into Asana"""

    def __init__(self):
        super().__init__(
            name="QuarterlyRoadmap",
            description="Import quarterly roadmap",
            version="1.0.0",
            category="asana"
        )
        self.project_email = None
        self.roadmap_file = None
        self.quarter = None

    def configure(self, **kwargs):
        """Configure the workflow"""
        self.project_email = kwargs.get('project_email')
        self.roadmap_file = kwargs.get('roadmap_file')
        self.quarter = kwargs.get('quarter')

        if not self.project_email:
            raise ValueError("project_email is required")
        if not self.roadmap_file:
            raise ValueError("roadmap_file is required")
        if not self.quarter:
            raise ValueError("quarter is required")

        logger.info(f"Configured roadmap import:")
        logger.info(f"  Quarter: {self.quarter}")
        logger.info(f"  Roadmap file: {self.roadmap_file}")
        logger.info(f"  Project email: {self.project_email}")

        return True

    def validate(self):
        """Validate configuration"""
        # Check file exists
        if not Path(self.roadmap_file).exists():
            raise FileNotFoundError(f"Roadmap file not found: {self.roadmap_file}")

        return True

    def execute(self):
        """Execute the workflow"""
        logger.info("=" * 60)
        logger.info(f"{self.quarter} ROADMAP IMPORT TO ASANA")
        logger.info("=" * 60)

        # Load roadmap from Excel
        logger.info(f"Loading roadmap from: {self.roadmap_file}")

        from openpyxl import load_workbook

        workbook = load_workbook(self.roadmap_file)
        sheet = workbook.active

        # Read header
        headers = [cell.value for cell in sheet[1] if cell.value]

        # Parse tasks
        tasks = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            task = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i] and value:
                    task[headers[i].lower().replace(' ', '_')] = str(value)

            # Add quarter tag
            if 'tags' in task:
                task['tags'] += f", {self.quarter}"
            else:
                task['tags'] = self.quarter

            # Ensure required fields
            if 'name' not in task and 'task_name' not in task:
                logger.warning(f"Row {row_idx} missing task name, skipping")
                continue

            tasks.append(task)

        logger.info(f"Loaded {len(tasks)} roadmap items from Excel")

        # Initialize helper
        helper = AsanaHelper(project_email=self.project_email)

        # Create tasks
        logger.info(f"Creating {len(tasks)} roadmap items in Asana...")
        result = helper.bulk_create_from_list(tasks, method='email')

        # Log results
        if result['status'] == 'success':
            data = result['data']
            logger.info(f"✓ Successfully created {data['created_count']} roadmap items")

            if data.get('failed_count', 0) > 0:
                logger.warning(f"✗ Failed to create {data['failed_count']} items")
                for failed in data['failed_tasks']:
                    logger.warning(f"  - {failed['name']}: {failed.get('error', 'Unknown error')}")

        logger.info("=" * 60)
        logger.info("ROADMAP IMPORT COMPLETE")
        logger.info(f"Check your Asana project for {self.quarter} tasks")
        logger.info("=" * 60)

        return result


def main():
    """Main entry point"""
    workflow = QuarterlyRoadmapWorkflow()

    # Example configuration
    result = workflow.run(
        project_email='q1-roadmap@mail.asana.com',
        roadmap_file='C:/Planning/Q1_2025_Roadmap.xlsx',
        quarter='Q1 2025'
    )

    print(f"\nResult: {result['message']}")


if __name__ == '__main__':
    main()
